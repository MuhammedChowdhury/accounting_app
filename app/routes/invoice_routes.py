from flask import Blueprint, request, jsonify, render_template, redirect, url_for, flash
from app.models import db, Invoice, Quote
import openpyxl
import os
from datetime import datetime
from flask import send_file

invoice_routes = Blueprint('invoice_routes', __name__)
script_dir = os.path.dirname(__file__)

### 🔹 Helper Function: Generate Invoice as Excel File ###
def create_excel_invoice(invoice_number, client_name, supplier_name, line_items, total_amount, due_date):
    """Creates a structured invoice file in Excel format."""
    file_name = f"Invoice_{invoice_number}.xlsx"
    excel_path = os.path.join(script_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ✅ Invoice Metadata (Top)
    ws["A1"] = "Invoice Number:"
    ws["B1"] = invoice_number
    ws["A2"] = "Date:"
    ws["B2"] = datetime.today().strftime('%Y-%m-%d')
    ws["A3"] = "Due Date:"
    ws["B3"] = due_date

    # ✅ Client Details (Left)
    ws["A5"] = "Client Name:"
    ws["B5"] = client_name

    # ✅ Supplier Details (Right)
    ws["D5"] = "Supplier Name:"
    ws["E5"] = supplier_name

    # ✅ Column Headers (Invoice Table)
    headers = ["Date", "Description", "Rate", "Quantity", "GST (10%)", "Total"]
    ws.append([""])  
    ws.append(headers)

    # ✅ Add Invoice Entries Dynamically
    for item in line_items:
        date = item.get("date")
        description = item.get("description")
        rate = float(item.get("rate", 0))
        quantity = int(item.get("quantity", 1))
        gst = rate * quantity * 0.10  
        total = rate * quantity  
        ws.append([date, description, rate, quantity, gst, total])

    # ✅ Total GST & Grand Total Calculation
    total_row = len(line_items) + 10
    ws[f"D{total_row}"] = "Total GST:"
    ws[f"E{total_row}"] = f"=SUM(E11:E{total_row - 1})"
    ws[f"D{total_row+1}"] = "Grand Total:"
    ws[f"E{total_row+1}"] = f"=SUM(F11:F{total_row - 1}) + SUM(E11:E{total_row - 1})"

    wb.save(excel_path)
    return file_name

### 🔹 Add Invoice ###
@invoice_routes.route('/add_invoice', methods=['POST'])
def add_invoice():
    """Receives invoice data & generates a separate Excel invoice file."""
    try:
        data = request.json  
        invoice_number = data.get("invoice_number")
        client_name = data.get("client_name")
        supplier_name = data.get("supplier_name")
        line_items = data.get("items", [])
        due_date = data.get("due_date")

        # Validate input
        if not invoice_number or not client_name or not supplier_name or not line_items or not due_date:
            return jsonify({"error": "Missing required fields"}), 400

        # Calculate total amount
        total_amount = sum(float(item.get("rate", 0)) * int(item.get("quantity", 1)) for item in line_items)

        # Save invoice in database
        new_invoice = Invoice(
            client_name=client_name,
            supplier_name=supplier_name,
            line_items=str(line_items),
            total_amount=total_amount,
            due_date=datetime.strptime(due_date, '%Y-%m-%d'),
            payment_status="Pending"
        )
        db.session.add(new_invoice)
        db.session.commit()

        file_name = create_excel_invoice(invoice_number, client_name, supplier_name, line_items, total_amount, due_date)
        
        return jsonify({"message": f"✅ Invoice '{file_name}' created!", "invoice": data}), 201

    except Exception as e:
        return jsonify({"error": f"Failed to add invoice: {str(e)}"}), 500

### 🔹 View Invoices ###
@invoice_routes.route('/view_invoices', methods=['GET'])
def view_invoices():
    try:
        invoices = Invoice.query.all()
        return render_template('view_invoices.html', invoices=invoices)
    except Exception as e:
        flash(f"Failed to fetch invoices: {str(e)}", "error")
        return redirect(url_for('invoice_routes.view_invoices'))

### 🔹 Convert Quote to Invoice ###
@invoice_routes.route('/convert_to_invoice/<int:quote_id>', methods=['POST'])
def convert_to_invoice(quote_id):
    try:
        quote = Quote.query.get(quote_id)
        if not quote:
            flash("Quote not found", "error")
            return redirect(url_for('invoice_routes.view_quotes'))

        new_invoice = Invoice(
            client_name="Customer Name",
            line_items=quote.item_name,
            total_amount=quote.quantity * quote.price,
            due_date=datetime.now().strftime('%Y-%m-%d'),
            payment_status="Pending"
        )
        db.session.add(new_invoice)
        db.session.commit()

        flash("Quote converted to invoice successfully!", "success")
        return redirect(url_for('invoice_routes.view_invoices'))
    except Exception as e:
        flash(f"Failed to convert quote: {str(e)}", "error")
        return redirect(url_for('invoice_routes.view_quotes'))

### 🔹 Mark Invoice as Paid ###
@invoice_routes.route('/mark_paid/<int:invoice_id>', methods=['POST'])
def mark_invoice_paid(invoice_id):
    try:
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            flash("Invoice not found", "error")
            return redirect(url_for('invoice_routes.view_invoices'))
        
        invoice.payment_status = "Paid"
        db.session.commit()
        flash("Invoice marked as paid!", "success")
        return redirect(url_for('invoice_routes.view_invoices'))
    except Exception as e:
        flash(f"Failed to update invoice: {str(e)}", "error")
        return redirect(url_for('invoice_routes.view_invoices'))

### 🔹 Delete Invoice ###
@invoice_routes.route('/delete_invoice/<int:invoice_id>', methods=['POST'])
def delete_invoice(invoice_id):
    try:
        invoice = Invoice.query.get(invoice_id)
        if not invoice:
            flash("Invoice not found", "error")
            return redirect(url_for('invoice_routes.view_invoices'))
        
        db.session.delete(invoice)
        db.session.commit()
        flash("Invoice deleted successfully!", "success")
        return redirect(url_for('invoice_routes.view_invoices'))
    except Exception as e:
        flash(f"Failed to delete invoice: {str(e)}", "error")
        return redirect(url_for('invoice_routes.view_invoices'))



@invoice_routes.route('/download_invoice/<invoice_number>', methods=['GET'])
def download_invoice(invoice_number):
    """Allows users to download a generated invoice file."""
    file_name = f"Invoice_{invoice_number}.xlsx"
    excel_path = os.path.join(script_dir, file_name)

    if not os.path.exists(excel_path):
        return jsonify({"error": "Invoice file not found!"}), 404

    return send_file(excel_path, as_attachment=True)
