import openpyxl
import os
from datetime import datetime

script_dir = os.path.dirname(__file__)  

def create_invoice(invoice_number, client_name):
    """Creates a separate invoice file for each transaction."""
    file_name = f"Invoice_{invoice_number}.xlsx"
    excel_path = os.path.join(script_dir, file_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"

    # ✅ Invoice Metadata (Top)
    ws["A1"] = "Invoice Number:"
    ws["B1"] = invoice_number
    ws["A2"] = "Date:"
    ws["B2"] = datetime.today().strftime('%Y-%m-%d')

    # ✅ Client Details
    ws["A4"] = "Client Name:"
    ws["B4"] = client_name
    ws["A5"] = "Address:"
    ws["B5"] = "Client Address"
    ws["A6"] = "Contact:"
    ws["B6"] = "Client Contact"

    # ✅ Column Headers (Invoice Table)
    headers = ["Date", "Description", "Rate", "Quantity", "GST (10%)", "Total"]
    ws.append([""])  # Blank row for spacing
    ws.append(headers)

    # ✅ Auto-filled formulas for GST & Total
    for row in range(9, 20):  # Pre-format rows for future invoices
        ws.cell(row=row, column=5).value = f"=0.10*D{row}"  # GST formula
        ws.cell(row=row, column=6).value = f"=C{row}*D{row}"  # Total formula

    # ✅ Grand Total Calculation
    ws["D21"] = "Total GST Sum:"
    ws["E21"] = "=SUM(E9:E19)"
    ws["D22"] = "Grand Total:"
    ws["E22"] = "=SUM(F9:F19) + SUM(E9:E19)"

    # Save as separate invoice file
    wb.save(excel_path)
    print(f"✅ Invoice {invoice_number} created: {excel_path}")

# Example: Create a new invoice
create_invoice("001", "Client ABC Pty Ltd")
